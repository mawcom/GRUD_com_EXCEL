from tkinter import * 
from tkinter import ttk
import openpyxl

#tratamendo de erro da planilha

try:
    wb=openpyxl.load_workbook('planilha_key.xlsx')
    ws=wb.active
    títulos=('Código','Produto','Vendedor','Quantidade')
    cont=1
    for c in títulos:
        celula=ws.cell(row=1,column=cont).value
        if c!=celula:
            ws.cell(row=1,column=cont,value=c)
            print(c)
            print(celula)
        cont+=1
    wb.save('planilha_key.xlsx')

except:
    títulos=('Código','Produto','Vendedor','Quantidade')
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title = 'planilha_key.xlsx'
    for num, c in enumerate(títulos):
        ws.cell(row=1,column=num+1,value=c)
    wb.save('planilha_key.xlsx')

#window

win = Tk()
win.configure(background='black')
part1 = Frame(win)
part1.place(relx=0.01, rely=0.01, relwidth=0.32, relheight=0.98)
part2 = Frame(win)
part2.place(relx=0.34, rely=0.01, relwidth=0.65, relheight=0.98)
win.geometry('650x250')

#Label's and Entry's
label_codigo = Label(part1, text='Código')
label_codigo.place(relx=0.01, rely=0.01)
cod = StringVar()
entry_codigo = Entry(part1, textvariable=cod)
entry_codigo.place(relx=0.01, rely=0.10, relwidth=0.60)

label_produto = Label(part1, text='Produto')
label_produto.place(relx=0.01, rely=0.20 )
pro = StringVar()
entry_produto = Entry(part1, textvariable=pro)
entry_produto.place(relx=0.01, rely=0.29, relwidth=0.60)

label_vendedor = Label(part1, text='Vendedor')
label_vendedor.place(relx=0.01,rely=0.39)
ven = StringVar()
entry_vendedor = Entry(part1, textvariable=ven)
entry_vendedor.place(relx=0.01, rely=0.48, relwidth=0.60)

label_quantidade = Label(part1, text='Quantidade')
label_quantidade.place(relx=0.01,rely=0.58)
qua = StringVar()
entry_quantidade= Entry(part1, textvariable=qua)
entry_quantidade.place(relx=0.01, rely=0.67, relwidth=0.60)

#def's

def limpar_entrys():
    entry_codigo.delete(0,END)
    entry_produto.delete(0,END)
    entry_quantidade.delete(0,END)
    entry_vendedor.delete(0,END)

def atuliazar_tree():
        for c in tv.get_children():
            tv.delete(c)
        for c in ws.iter_rows(min_row=2, values_only=True):
            tv.insert('',END,values=c)

def ney(alguma_coisa_man):
    """seleção de dados e aquisiçao de index para apagar e editar"""
    #vai tomar nomar no cu values dessa maldita 
    #função .item que não retorna a lista com todos valores em string, fazer gambiarra pra inferno
    for item in tv.selection():
        print(rowselect())
        bagulho=tv.item(item)
        limpar_entrys()
        codiguin=str(bagulho['values'][0])[1:]
        entry_codigo.insert(0,codiguin)
        entry_produto.insert(0,bagulho['values'][1])
        entry_quantidade.insert(0,bagulho['values'][3])
        entry_vendedor.insert(0,bagulho['values'][2])

def rowselect():
    """retorna o index da row que foi selecionada no banco de dados """
    for num, c in enumerate(tv.get_children()):
        tamanho_tree=num
    tamanho_tree=tamanho_tree+1
    if tamanho_tree==len(ws['A'])-1:
        bagulho=tv.selection()[0]
        cont=2
        for c in tv.get_children():
            if c==bagulho:
                break
            cont+=1
            print(cont)
    else:
        for item in tv.selection():
            bagulho=tv.item(item)
            limpar_entrys()
            cont=2
            bagulho_tratado=[]
            for c in bagulho['values']:
                bagulho_tratado.append(str(c)) 
            bagulho_tratado=tuple(bagulho_tratado)
        for c in ws.iter_rows(min_row=2, values_only=True):
            if c==bagulho_tratado:
                print(bagulho_tratado)
                print(ws.cell(row=cont,column=1).value)
                break
            cont+=1
    
    return cont

def salvar_adicionar():
    dados = (str(f'#{cod.get()}'),str(pro.get()),str(ven.get()),str(qua.get()))
    dados_lista=['#','-','-','-']
    for num,c in enumerate(dados):
        if c!='':
            dados_lista.pop(num)
            dados_lista.insert(num,c)
    dados=tuple(dados_lista)
    lfd = len(ws['A'])+1
    for num, c in enumerate(dados):
        ws.cell(row=lfd, column=num+1, value=c)
    tv.insert('',END,values=dados)
    wb.save('planilha_key.xlsx')
    limpar_entrys()


def apaga_fuc():
    ws.delete_rows(rowselect())
    wb.save('planilha_key.xlsx')
    atuliazar_tree()
    limpar_entrys()

    
def editar():
    cel=tv.selection()
    dados = (str(f'#{cod.get()}'),str(pro.get()),str(ven.get()),str(qua.get()))
    linha=(rowselect())
    for num, c in enumerate(dados):
        ws.cell(row=linha, column=num+1, value=c)
    tv.insert('',linha-1, values=dados)
    tv.delete(cel)
    wb.save('planilha_key.xlsx')
    limpar_entrys()


def buscador():
    dados = (str(f'{cod.get()}'),str(pro.get()),str(ven.get()),str(qua.get()))
    controle_mestre=0
    valor=''
    posição=0
    for num,c in enumerate(dados):
        if c!='':
            posição=num
            controle_mestre=1
            valor=c
            break
    if controle_mestre==1:
        if posição==0:
            for c in tv.get_children():
                if tv.item(c)['values'][posição]==f'#{valor}':
                    print(tv.item(c)['values'][posição])
                else:
                    tv.delete(c)
        else:
            for c in tv.get_children():
                if str(tv.item(c)['values'][posição])==str(valor):
                    print(tv.item(c)['values'][posição])
                else:
                    tv.delete(c)
    if controle_mestre==0:
        atuliazar_tree()
        
#botões

button_add = Button(part1, text='Adicionar', command= lambda: salvar_adicionar())
button_add.place(relx=0.01, rely=0.77, relwidth=0.49,relheight=0.11)

button_a = Button(part1, text='Apagar', command= lambda: apaga_fuc())
button_a.place(relx=0.01, rely=0.88, relwidth= 0.49,relheight=0.11)

button_e = Button(part1, text='Editar', command= lambda:editar())
button_e.place(relx=0.50, rely=0.77, relwidth= 0.49,relheight=0.11)

button_b = Button(part1, text='Buscar',command= lambda:buscador())
button_b.place(relx=0.50, rely=0.88, relwidth= 0.49,relheight=0.11)

#tree view
tittle_kyes = ('codigo','produto', 'vendedor', 'quantidade')
tv = ttk.Treeview(part2, columns=tittle_kyes, show='headings')

tv.heading('codigo', text='Código')
tv.heading('produto', text='Produto')
tv.heading('vendedor', text='Vendedor')
tv.heading('quantidade', text='Quantidade')

tv.column('codigo', width=94)
tv.column('produto', width=94)
tv.column('vendedor', width=94)
tv.column('quantidade', width=94)
tv.place(relx=0,rely=0,relheight=1,relwidth=1)
barra = ttk.Scrollbar(part2, orient=VERTICAL, command=tv.yview)

barra_rolagem = ttk.Scrollbar(part2, orient=VERTICAL, command=tv.yview)
tv.configure(yscrollcommand=barra_rolagem.set)
barra_rolagem.place(relx=0.96,rely=0.01,relheight=0.99)

for c in ws.iter_rows(min_row=2, values_only=True):
    tv.insert('',END,values=c)

tv.bind('<<TreeviewSelect>>', ney)

win.mainloop()

